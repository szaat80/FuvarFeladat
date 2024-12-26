from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QFrame,
    QPushButton, QLabel, QLineEdit, QDateEdit, QTimeEdit,
    QComboBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QApplication, QMenuBar, QMenu, QFileDialog, QMessageBox,
    QDialog
)
from PySide6.QtCore import Qt, QTime, QDate
from PySide6.QtGui import QFont, QColor
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtWidgets import QFormLayout, QTabWidget, QSpinBox
import sys
import json
import calendar
import sqlite3
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

class MenuBar(QMenuBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.createFileMenu()
        self.createDatabaseMenu()

    def createFileMenu(self):
        fileMenu = self.addMenu("Fájl")
        fileMenu.addAction("Excel megnyitása").triggered.connect(self.parent().openExcel)
        fileMenu.addAction("Mentés").triggered.connect(self.parent().saveToExcel)
        fileMenu.addAction("Nyomtatás").triggered.connect(self.parent().printData)
        fileMenu.addAction("Kilépés").triggered.connect(self.parent().close)

    def createDatabaseMenu(self):
        dbMenu = self.addMenu("Adatbázis")
        dbMenu.addAction("Törzsadatok kezelése").triggered.connect(self.parent().openDatabaseManager)

class FuvarAdminApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initDatabase()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Fuvar Adminisztráció")
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        self.setMenuBar(MenuBar(self))
        self.setupStyles()
        self.setupTopFrame(main_layout)
        self.setupBottomFrame(main_layout)
        self.setupButtons(main_layout)
        main_widget.setLayout(main_layout)
        self.showMaximized()

    def initDatabase(self):
        self.conn = sqlite3.connect('fuvarok.db')
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS factories (
                id INTEGER PRIMARY KEY,
                nev TEXT,
                fuvardij INTEGER
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS addresses (
                id INTEGER PRIMARY KEY,
                cim TEXT,
                ar INTEGER
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS zones (
                id INTEGER PRIMARY KEY,
                nev TEXT,
                alapdij INTEGER
            )
        ''')
        self.conn.commit()
        
        cursor.execute("SELECT COUNT(*) FROM factories")
        count = cursor.fetchone()[0]
        
        if count == 0:
            factories = [
                ('CATL', 5000),
                ('BMW', 6000),
                ('Gyár 3', 5500)
            ]
            cursor.executemany("INSERT INTO factories (nev, fuvardij) VALUES (?, ?)", factories)
            self.conn.commit()

    def setupStyles(self):
        self.styles = {
            'main_frame': """
                QFrame {
                    background-color: #2d2d2d;
                    border: 3px solid #3e3e3e;
                    border-radius: 15px;
                    margin: 5px;
                }
            """,
            'sub_frame': """
                QFrame {
                    background-color: #d9d9d9;
                    border: 2px solid #4d4d4d;
                    border-radius: 15px;
                    margin: 5px;
                }
            """,
            'table_frame': """
                QFrame {
                    background-color: #d9d9d9;
                    border: 3px solid #ff2800;
                    border-radius: 15px;
                    padding: 10px;
                }
            """,
            'label': """
                QLabel {
                    color: black;
                    font-size: 14px;
                    font-weight: bold;
                    padding: 5px;
                }
            """,
            'input': """
                QLineEdit, QDateEdit, QTimeEdit, QComboBox {
                    background-color: white; 
                    padding: 5px;
                    border: 2px solid #a0a0a0;
                    border-radius: 5px;
                    color: black;
                    min-width: 150px;
                    max-width: 150px;
                    min-height: 30px;
                    max-height: 30px;
                    font-size: 14px;
                }
            """,
            'button': """
                QPushButton {
                    background-color: #4a90e2;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    min-width: 120px;
                }
                QPushButton:hover {
                    background-color: #357abd;
                }
            """
        }

    def setupTopFrame(self, main_layout):
        top_frame = QFrame()
        top_frame.setStyleSheet(self.styles['main_frame'])
        top_layout = QHBoxLayout()
        
        # Bal panel
        left_panel = QFrame()
        left_panel.setStyleSheet(self.styles['sub_frame'])
        left_layout = QVBoxLayout()
        
        # Fix szélesség minden beviteli mezőhöz
        input_width = 150
        
        # Dátum mező
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setFixedWidth(input_width)
        left_layout.addLayout(self.createInputGroup("Dátum:", self.date_edit))
        
        # Idő mezők
        self.start_time = QTimeEdit()
        self.start_time.setButtonSymbols(QTimeEdit.NoButtons)
        self.start_time.setFixedWidth(input_width)
        self.end_time = QTimeEdit()
        self.end_time.setButtonSymbols(QTimeEdit.NoButtons)
        self.end_time.setFixedWidth(input_width)
        left_layout.addLayout(self.createInputGroup("Kezdés:", self.start_time))
        left_layout.addLayout(self.createInputGroup("Végzés:", self.end_time))
        
        # Munka típus
        self.type_combo = QComboBox()
        self.type_combo.addItems(["Sima munkanap", "Műhely nap", "Szabadság", "Betegszabadság (TP)"])
        self.type_combo.setFixedWidth(input_width)
        left_layout.addLayout(self.createInputGroup("Munka típusa:", self.type_combo))
        
        left_layout.addStretch()
        left_panel.setLayout(left_layout)
        
        # Jobb panel
        right_panel = QFrame()
        right_panel.setStyleSheet(self.styles['sub_frame'])
        right_layout = QVBoxLayout()
        
        # Kilométer sáv
        self.km_combo = QComboBox()
        self.km_combo.addItems([f"Övezet {i}-{i+5}" for i in range(0, 50, 5)])
        self.km_combo.setFixedWidth(input_width)
        right_layout.addLayout(self.createInputGroup("Kilométer sáv:", self.km_combo))
        
        # Gyár választó
        self.factory_combo = QComboBox()
        self.factory_combo.setFixedWidth(input_width)
        self.loadFactories()
        right_layout.addLayout(self.createInputGroup("Gyár:", self.factory_combo))
        
        # Cím és szállítószám
        self.address_input = QLineEdit()
        self.address_input.setFixedWidth(input_width)
        right_layout.addLayout(self.createInputGroup("Cím:", self.address_input))
        
        self.delivery_input = QLineEdit()
        self.delivery_input.setFixedWidth(input_width)
        right_layout.addLayout(self.createInputGroup("Szállítószám:", self.delivery_input))
        
        # M3 bevitel
        self.m3_input = QLineEdit()
        self.m3_input.setFixedWidth(input_width)
        self.m3_input.returnPressed.connect(self.handleM3Input)
        self.m3_sum = QLabel("(0)")
        
        m3_layout = QHBoxLayout()
        m3_label = QLabel("M3:")
        m3_label.setStyleSheet(self.styles['label'])
        m3_layout.addWidget(m3_label)
        m3_layout.addWidget(self.m3_input)
        m3_layout.addWidget(self.m3_sum)
        m3_layout.addStretch()
        
        right_layout.addLayout(m3_layout)
        right_layout.addStretch()
        right_panel.setLayout(right_layout)
        
        top_layout.addWidget(left_panel)
        top_layout.addWidget(right_panel)
        top_frame.setLayout(top_layout)
        main_layout.addWidget(top_frame)

    def setupBottomFrame(self, main_layout):
        bottom_frame = QFrame()
        bottom_frame.setStyleSheet(self.styles['main_frame'])
        button_layout = QHBoxLayout()

        # Munkaórák táblázat
        work_frame = QFrame()
        work_frame.setStyleSheet(self.styles['table_frame'])
        work_layout = QVBoxLayout()

        # Munkaórák táblázat fejlécének beállítása
        self.work_table = QTableWidget()
        self.work_table.setColumnCount(7)
        work_headers = ["Dátum", "Nap", "Munka KB", "Munka BF", 
                       "Ledolgozott óra", "Műhely KB", "Műhely BF"]
    
        header = self.work_table.horizontalHeader()
        header.setVisible(True)
        header.setMinimumHeight(40)
        for i, title in enumerate(work_headers):
            item = QTableWidgetItem(title)
            item.setTextAlignment(Qt.AlignCenter)
            self.work_table.setHorizontalHeaderItem(i, item)
            self.work_table.setColumnWidth(i, 150)
    
        # Fuvar táblázat fejlécének beállítása
        self.delivery_table = QTableWidget()
        self.delivery_table.setColumnCount(10)
        delivery_headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)] + ["Összeg"]
    
        header = self.delivery_table.horizontalHeader()
        header.setVisible(True)
        header.setMinimumHeight(40)
        for i, title in enumerate(delivery_headers):
            item = QTableWidgetItem(title)
            item.setTextAlignment(Qt.AlignCenter)
            self.delivery_table.setHorizontalHeaderItem(i, item)
            self.delivery_table.setColumnWidth(i, 150)
    
        # Fejléc formázása
        header = self.work_table.horizontalHeader()
        header.setFixedHeight(35)  # Fejléc magasság
        header.setDefaultAlignment(Qt.AlignCenter)  # Középre igazítás
        header.setStyleSheet("""
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 5px;
                border: 1px solid black;
                font-size: 14px;
                font-weight: bold;
            }
        """)
    
        # Oszlopszélességek beállítása
        for i in range(self.work_table.columnCount()):
            self.work_table.setColumnWidth(i, 150)

        work_layout.addWidget(self.work_table)
        work_frame.setLayout(work_layout)
        button_layout.addWidget(work_frame)

        # Fuvar táblázat
        delivery_frame = QFrame()
        delivery_frame.setStyleSheet(self.styles['table_frame'])
        delivery_layout = QVBoxLayout()

        self.delivery_table = QTableWidget()
        self.delivery_table.setColumnCount(10)
        delivery_headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)] + ["Összeg"]
    
        header = self.delivery_table.horizontalHeader()
        header.setVisible(True)
        header.setMinimumHeight(40)
        for i, title in enumerate(delivery_headers):
            item = QTableWidgetItem(title)
            item.setTextAlignment(Qt.AlignCenter)
            self.delivery_table.setHorizontalHeaderItem(i, item)
            self.delivery_table.setColumnWidth(i, 150)
    
        # Fejlécek beállítása
        delivery_headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)] + ["Összeg"]
        self.delivery_table.setHorizontalHeaderLabels(delivery_headers)
    
        # Fejléc formázása
        header = self.delivery_table.horizontalHeader()
        header.setFixedHeight(35)
        header.setDefaultAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 5px;
                border: 1px solid black;
                font-size: 14px;
                font-weight: bold;
            }
        """)
    
        # Oszlopszélességek beállítása
        for i in range(self.delivery_table.columnCount()):
            self.delivery_table.setColumnWidth(i, 150)

        delivery_layout.addWidget(self.delivery_table)
        delivery_frame.setLayout(delivery_layout)
        button_layout.addWidget(delivery_frame)

        bottom_frame.setLayout(button_layout)
        main_layout.addWidget(bottom_frame)

        self.setupTableRows()
        self.setupTableStyles()

    def setupTableStyles(self):
        table_style = """
            QTableWidget {
                background-color: white;
                border: 2px solid #ff2800;
                color: black;
                font-size: 14px;
                gridline-color: black;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                color: black;
                font-weight: bold;
                border: 1px solid black;
                padding: 8px;
                font-size: 14px;
                min-height: 35px;
                font-family: Arial;
            }
            QTableWidget::item {
                padding: 8px;
                min-height: 30px;
                border-bottom: 1px solid #ddd;
            }
            QScrollBar {
                background: white;
            }
        """
        self.work_table.setStyleSheet(table_style)
        self.delivery_table.setStyleSheet(table_style)

    def setupTableRows(self):
        current_date = QDate.currentDate()
        first_day = QDate(current_date.year(), current_date.month(), 1)
        days_in_month = first_day.daysInMonth()
        
        # Magyar napnevek
        day_names = ['Hétfő', 'Kedd', 'Szerda', 'Csütörtök', 'Péntek', 'Szombat', 'Vasárnap']
        
        for table in [self.work_table, self.delivery_table]:
            table.setRowCount(0)  # Töröljük a meglévő sorokat
            
            for i in range(days_in_month):
                current_day = first_day.addDays(i)
                table.insertRow(i)
                
                # Dátum beállítása
                date_item = QTableWidgetItem(current_day.toString('yyyy-MM-dd'))
                date_item.setTextAlignment(Qt.AlignCenter)
                table.setItem(i, 0, date_item)
                
                if table == self.work_table:
                    # Nap neve
                    day_name = day_names[current_day.dayOfWeek() - 1]
                    day_item = QTableWidgetItem(day_name)
                    day_item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(i, 1, day_item)

    def createInputGroup(self, label_text, widget):
        layout = QHBoxLayout()
        label = QLabel(label_text)
        label.setStyleSheet(self.styles['label'])
        layout.addWidget(label)
        widget.setStyleSheet(self.styles['input'])
        layout.addWidget(widget)
        layout.addStretch()
        return layout

    def setupButtons(self, main_layout):
        button_frame = QFrame()
        button_layout = QHBoxLayout()
        
        buttons = [
            ("Munkaórák Mentése", self.saveWorkHours),
            ("Fuvar Adatok Mentése", self.saveDeliveryData),
            ("Excel Exportálás", self.exportToExcel),
            ("Kilépés", self.close)
        ]
        
        for text, callback in buttons:
            btn = QPushButton(text)
            btn.clicked.connect(callback)
            btn.setStyleSheet(self.styles['button'])
            button_layout.addWidget(btn)
        
        button_frame.setLayout(button_layout)
        main_layout.addWidget(button_frame)

    def loadFactories(self):
        self.factory_combo.clear()
        cursor = self.conn.cursor()
        cursor.execute("SELECT nev FROM factories")
        factories = cursor.fetchall()
        self.factory_combo.addItems([factory[0] for factory in factories])

    def handleM3Input(self):
        text = self.m3_input.text().strip()
        try:
            # Vesszőt pontra cserélünk
            text = text.replace(',', '.')
            # Ellenőrizzük, hogy van-e érték
            if not text:
                return
            
            value = float(text)
            if not hasattr(self, 'm3_values'):
                self.m3_values = []
            self.m3_values.append(value)
        
            # Formázott megjelenítés
            display_values = []
            for v in self.m3_values:
                display_values.append(f"{v:.1f}")
        
            display_text = ' + '.join(display_values)
            total = sum(self.m3_values)
            sum_text = f"({display_text}) ({total:.1f})"
            self.m3_sum.setText(sum_text)
        
            self.updateDeliveryTable(sum_text)
            self.m3_input.clear()
        except ValueError:
            QMessageBox.warning(self, "Hiba", "Kérem számot adjon meg (pl.: 6.0 vagy 6,0)")

    def updateDeliveryTable(self, sum_text):
        date_text = self.date_edit.date().toString('yyyy-MM-dd')
        zone_text = self.km_combo.currentText()
        zone_col = self.getZoneColumn(zone_text)
        
        if zone_col > 0:
            for row in range(self.delivery_table.rowCount()):
                if self.delivery_table.item(row, 0).text() == date_text:
                    self.delivery_table.setItem(row, zone_col, QTableWidgetItem(sum_text))

    def getZoneColumn(self, zone_text):
        try:
            start_km = int(zone_text.split(' ')[1].split('-')[0])
            column = (start_km // 5) + 1
            return column
        except:
            return 0

    def saveWorkHours(self):
        try:
            data = {
                'date': self.date_edit.date().toString('yyyy-MM-dd'),
                'start_time': self.start_time.time().toString('HH:mm'),
                'end_time': self.end_time.time().toString('HH:mm'),
                'type': self.type_combo.currentText()
            }
        
            # Táblázat frissítése
            date_text = data['date']
            start_text = data['start_time']
            end_text = data['end_time']
        
            # Táblázat frissítése
            for row in range(self.work_table.rowCount()):
                if self.work_table.item(row, 0).text() == date_text:
                    self.work_table.setItem(row, 2, QTableWidgetItem(start_text))
                    self.work_table.setItem(row, 3, QTableWidgetItem(end_text))
                    
                    # Ledolgozott órák számítása
                    start = datetime.strptime(start_text, '%H:%M')
                    end = datetime.strptime(end_text, '%H:%M')
                    hours = (end - start).seconds / 3600
                    self.work_table.setItem(row, 4, QTableWidgetItem(f"{hours:.2f}"))

            # Adatok mentése JSON fájlba
            with open('work_hours.json', 'a', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
                f.write('\n')
            
            QMessageBox.information(self, "Siker", "Munkaórák mentve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Hiba történt: {str(e)}")

    def saveDeliveryData(self):
        try:
            data = {
                'date': self.date_edit.date().toString('yyyy-MM-dd'),
                'km_range': self.km_combo.currentText(),
                'factory': self.factory_combo.currentText(),
                'address': self.address_input.text(),
                'delivery_number': self.delivery_input.text(),
                'm3_values': self.m3_values if hasattr(self, 'm3_values') else []
            }
        
            # Táblázat frissítése
            date_text = data['date']
            km_range = data['km_range']
            m3_values = data['m3_values']
        
            # Övezet oszlop meghatározása
            zone_col = self.getZoneColumn(km_range)
        
            if zone_col > 0:
                for row in range(self.delivery_table.rowCount()):
                    if self.delivery_table.item(row, 0).text() == date_text:
                        # M3 értékek összege
                        m3_sum = sum(m3_values) if m3_values else 0
                    
                        # Meglévő érték ellenőrzése
                        current_item = self.delivery_table.item(row, zone_col)
                        current_value = float(current_item.text()) if current_item and current_item.text() else 0
                    
                        # Új érték beállítása
                        new_value = current_value + m3_sum
                        self.delivery_table.setItem(row, zone_col, QTableWidgetItem(f"{new_value:.1f}"))

            # Adatok mentése JSON fájlba
            with open('delivery_data.json', 'a', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
                f.write('\n')
        
            # M3 értékek törlése a következő bevitelhez
            self.m3_values = []
            self.m3_input.clear()
            self.m3_sum.setText("(0)")
        
            QMessageBox.information(self, "Siker", "Fuvar adatok mentve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Hiba történt: {str(e)}")

    def exportToExcel(self):
        try:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Munkaórák"
            
            # Fejlécek
            headers = ["Dátum", "Nap", "Kezdés", "Végzés", "Munka típusa"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Adatok mentése
            for row in range(self.work_table.rowCount()):
                for col in range(len(headers)):
                    item = self.work_table.item(row, col)
                    if item:
                        ws1.cell(row=row+2, column=col+1, value=item.text())
            
            wb.save('munka_nyilvantartas.xlsx')
            QMessageBox.information(self, "Siker", "Excel fájl mentve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Mentési hiba: {str(e)}")

    def saveToExcel(self):
        try:
            wb = Workbook()
        
            # Munkaórák munkalap
            ws1 = wb.active
            ws1.title = "Munkaórák"
            work_headers = ["Dátum", "Nap", "Munka KB", "Munka BF", 
                           "Ledolgozott óra", "Műhely KB", "Műhely BF"]
        
            # Fejlécek formázása az első munkalapon
            for col, header in enumerate(work_headers, 1):
                cell = ws1.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Munkaóra adatok mentése
            for row in range(self.work_table.rowCount()):
                for col in range(len(work_headers)):
                    item = self.work_table.item(row, col)
                    if item:
                        ws1.cell(row=row+2, column=col+1, value=item.text())
        
            # Fuvar adatok munkalap
            ws2 = wb.create_sheet(title="Fuvar adatok")
            delivery_headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)] + ["Összeg"]
        
            # Fejlécek formázása a második munkalapon
            for col, header in enumerate(delivery_headers, 1):
                cell = ws2.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Fuvar adatok mentése
            for row in range(self.delivery_table.rowCount()):
                for col in range(len(delivery_headers)):
                    item = self.delivery_table.item(row, col)
                    if item:
                        ws2.cell(row=row+2, column=col+1, value=item.text())
        
            # Excel fájl mentése
            wb.save('munka_nyilvantartas.xlsx')
            QMessageBox.information(self, "Siker", "Excel fájl mentve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Mentési hiba: {str(e)}")

    def openDatabaseManager(self):
        dbManager = DatabaseManager(self)
        dbManager.exec_()

    def openExcel(self):
        try:
            file_name, _ = QFileDialog.getOpenFileName(
                self, "Excel fájl megnyitása", "", "Excel files (*.xlsx *.xls)"
            )
            if file_name:
                wb = load_workbook(file_name)
                self.loadDataFromExcel(wb)
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Fájl megnyitási hiba: {str(e)}")

    def loadDataFromExcel(self, workbook):
        try:
            # Munkaórák betöltése
            if "Munkaórák" in workbook.sheetnames:
                ws = workbook["Munkaórák"]
                for row in range(2, ws.max_row + 1):
                    date = ws.cell(row=row, column=1).value
                    if date:
                        self.loadWorkHoursRow(date, ws, row)
            
            # Fuvar adatok betöltése
            if "Fuvar adatok" in workbook.sheetnames:
                ws = workbook["Fuvar adatok"]
                for row in range(2, ws.max_row + 1):
                    date = ws.cell(row=row, column=1).value
                    if date:
                        self.loadDeliveryRow(date, ws, row)
                        
            QMessageBox.information(self, "Siker", "Excel adatok betöltve!")
        except Exception as e:
            QMessageBox.warning(self, "Hiba", f"Adatok betöltési hiba: {str(e)}")

    def loadWorkHoursRow(self, date, ws, row):
        start_time = ws.cell(row=row, column=2).value
        end_time = ws.cell(row=row, column=3).value
        work_type = ws.cell(row=row, column=4).value
        
        for table_row in range(self.work_table.rowCount()):
            if self.work_table.item(table_row, 0).text() == str(date):
                if start_time:
                    self.work_table.setItem(table_row, 2, QTableWidgetItem(str(start_time)))
                if end_time:
                    self.work_table.setItem(table_row, 3, QTableWidgetItem(str(end_time)))
                if work_type:
                    self.work_table.setItem(table_row, 4, QTableWidgetItem(str(work_type)))

    def loadDeliveryRow(self, date, ws, row):
        for col in range(1, self.delivery_table.columnCount()):
            value = ws.cell(row=row, column=col+1).value
            if value is not None:
                for table_row in range(self.delivery_table.rowCount()):
                    if self.delivery_table.item(table_row, 0).text() == str(date):
                        self.delivery_table.setItem(table_row, col, QTableWidgetItem(str(value)))

    def printData(self):
        dialog = QPrintDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            printer = dialog.printer()
            # TODO: Nyomtatási logika implementálása

class DatabaseManager(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.conn = sqlite3.connect('fuvarok.db')
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Törzsadat Kezelő")
        self.setMinimumWidth(800)
        self.setMinimumHeight(600)
        
        layout = QVBoxLayout()
        
        # Tab widget létrehozása
        tabs = QTabWidget()
        tabs.addTab(self.createFactoriesTab(), "Gyárak")
        tabs.addTab(self.createAddressesTab(), "Címek")
        tabs.addTab(self.createZonesTab(), "Övezetek")
        layout.addWidget(tabs)
        
        self.setLayout(layout)

    def createFactoriesTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Adatbeviteli mezők
        form_layout = QFormLayout()
        self.factory_name = QLineEdit()
        self.factory_price = QSpinBox()
        self.factory_price.setRange(0, 1000000)
        form_layout.addRow("Gyár neve:", self.factory_name)
        form_layout.addRow("Fuvardíj:", self.factory_price)
        
        # Gombok
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Hozzáadás")
        add_btn.clicked.connect(self.addFactory)
        delete_btn = QPushButton("Törlés")
        delete_btn.clicked.connect(self.deleteFactory)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(delete_btn)
        
        # Táblázat
        self.factory_table = QTableWidget()
        self.factory_table.setColumnCount(3)
        self.factory_table.setHorizontalHeaderLabels(["ID", "Név", "Fuvardíj"])
        
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.factory_table)
        widget.setLayout(layout)
        
        self.loadFactories()
        return widget

    def createAddressesTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        self.address = QLineEdit()
        self.address_price = QSpinBox()
        self.address_price.setRange(0, 1000000)
        form_layout.addRow("Cím:", self.address)
        form_layout.addRow("Egyedi ár:", self.address_price)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Hozzáadás")
        add_btn.clicked.connect(self.addAddress)
        delete_btn = QPushButton("Törlés")
        delete_btn.clicked.connect(self.deleteAddress)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(delete_btn)
        
        self.address_table = QTableWidget()
        self.address_table.setColumnCount(3)
        self.address_table.setHorizontalHeaderLabels(["ID", "Cím", "Ár"])
        
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.address_table)
        widget.setLayout(layout)
        
        self.loadAddresses()
        return widget

    def createZonesTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        self.zone_name = QLineEdit()
        self.zone_price = QSpinBox()
        self.zone_price.setRange(0, 1000000)
        form_layout.addRow("Övezet:", self.zone_name)
        form_layout.addRow("Alapdíj:", self.zone_price)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Hozzáadás")
        add_btn.clicked.connect(self.addZone)
        delete_btn = QPushButton("Törlés")
        delete_btn.clicked.connect(self.deleteZone)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(delete_btn)
        
        self.zone_table = QTableWidget()
        self.zone_table.setColumnCount(3)
        self.zone_table.setHorizontalHeaderLabels(["ID", "Név", "Alapdíj"])
        
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.zone_table)
        widget.setLayout(layout)
        
        self.loadZones()
        return widget

    def loadFactories(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM factories")
        self.factory_table.setRowCount(0)
        for row, data in enumerate(cursor.fetchall()):
            self.factory_table.insertRow(row)
            for col, item in enumerate(data):
                self.factory_table.setItem(row, col, QTableWidgetItem(str(item)))

    def loadAddresses(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM addresses")
        self.address_table.setRowCount(0)
        for row, data in enumerate(cursor.fetchall()):
            self.address_table.insertRow(row)
            for col, item in enumerate(data):
                self.address_table.setItem(row, col, QTableWidgetItem(str(item)))

    def loadZones(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM zones")
        self.zone_table.setRowCount(0)
        for row, data in enumerate(cursor.fetchall()):
            self.zone_table.insertRow(row)
            for col, item in enumerate(data):
                self.zone_table.setItem(row, col, QTableWidgetItem(str(item)))

    def addFactory(self):
        name = self.factory_name.text()
        price = self.factory_price.value()
        if name and price:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO factories (nev, fuvardij) VALUES (?, ?)", (name, price))
            self.conn.commit()
            self.loadFactories()
            self.factory_name.clear()
            self.factory_price.setValue(0)

    def addAddress(self):
        address = self.address.text()
        price = self.address_price.value()
        if address and price:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO addresses (cim, ar) VALUES (?, ?)", (address, price))
            self.conn.commit()
            self.loadAddresses()
            self.address.clear()
            self.address_price.setValue(0)

    def addZone(self):
        name = self.zone_name.text()
        price = self.zone_price.value()
        if name and price:
            cursor = self.conn.cursor()
            cursor.execute("INSERT INTO zones (nev, alapdij) VALUES (?, ?)", (name, price))
            self.conn.commit()
            self.loadZones()
            self.zone_name.clear()
            self.zone_price.setValue(0)

    def deleteFactory(self):
        selected = self.factory_table.selectedItems()
        if selected:
            row = selected[0].row()
            id_item = self.factory_table.item(row, 0)
            if id_item:
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM factories WHERE id=?", (id_item.text(),))
                self.conn.commit()
                self.loadFactories()

    def deleteAddress(self):
        selected = self.address_table.selectedItems()
        if selected:
            row = selected[0].row()
            id_item = self.address_table.item(row, 0)
            if id_item:
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM addresses WHERE id=?", (id_item.text(),))
                self.conn.commit()
                self.loadAddresses()

    def deleteZone(self):
        selected = self.zone_table.selectedItems()
        if selected:
            row = selected[0].row()
            id_item = self.zone_table.item(row, 0)
            if id_item:
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM zones WHERE id=?", (id_item.text(),))
                self.conn.commit()
                self.loadZones()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FuvarAdminApp()
    window.show()
    sys.exit(app.exec())  # Eltávolítottuk az aláhúzást